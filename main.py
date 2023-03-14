from openpyxl import load_workbook as WB
import pyexcel
from pyexcel_xls import save_data
from fileinput import filename
import pandas as pd


class work_doc_csv(object): #(класс записи и считывания данных)
    def __init__(self, filename= './Таблица артикулов.xlsx', articles = [] ):
        self.filename = filename
        self.articles = articles

    def write_File(self, count_of_articles = 3, count_with_hyphen = False, list_of_articles = []):
        a = ''
        if count_with_hyphen == True:
            a = '-'
        name_of_title = str(count_of_articles) + a
        name_of_doc = name_of_title + '.xls'
        save_data(name_of_doc, list_of_articles)


    def open_File(self): #(Метод в котором описывается открытие файла, и считывание 1-го ряда символов)
        wb = WB(self.filename)
        tit: object = wb.sheetnames
        work_doc = wb[tit[0]]
        wb.columns = 'A'
        for i in range(1, work_doc.max_row): # 2-я цыфра - это цифра длинны всего массива в эксель документе
            self.articles.append(work_doc.cell(row=i, column=1).value)
        articles = self.articles
        return articles


class sorted_article_csv(object):
    def __init__(self):
        pass

    def sorted_article(self, list = []):
        new_article = []
        new_article_biggest3 = []
        for num in range(0, len(list)):
            write = True
            for i in range(0, len(list[num])):
                if list[num][i] != "1":
                    if list[num][i] != "2":
                        if list[num][i] != "3":
                            if list[num][i] != "4":
                                if list[num][i] != "5":
                                    if list[num][i] != "6":
                                        if list[num][i] != "7":
                                            if list[num][i] != "8":
                                                if list[num][i] != "9":
                                                    if list[num][i] != "-":
                                                        if list[num][i] != "0":
                                                            write = False
                                                            break

            if write == True:
                new_article.append(list[num])
        new_article.sort()
        for key in range(0, len(new_article)):
            if len(new_article[key]) >= 3:
                new_article_biggest3.append(new_article[key])
        return new_article_biggest3

    def view_all_article(self, sorted_list=[], size_of_article=4, count_with_hyphen=False):
        return_article = []
        if count_with_hyphen==False:
            for art in sorted_list:
                if len(art) == size_of_article:
                    return_article.append(art)
        if count_with_hyphen==True:
            for art in sorted_list:
                if len(art) > size_of_article :
                    if art[size_of_article] == '-':
                        return_article.append(art[0:size_of_article])
        return return_article


class function_with_article(object):
    def __init__(self, first_list):
        self.first_list = first_list

    def delete_one_article_from_another(self, deleted_list_of_article=[]):
        new_article_list = []
        for num_first_article in range(0, len(self.first_list)):
            load_article = True
            for num_deleted_article in range(0, len(deleted_list_of_article)):
                if self.first_list[num_first_article] == deleted_list_of_article[num_deleted_article]:
                    load_article = False
            if load_article == True:
                new_article_list.append(self.first_list[num_first_article])
        return new_article_list


def make_all_visible_articles(count_of_articles):
    all_usible_article = []  # список всех коунтеров
    if count_of_articles == 1:
        for i in range(1, 9):
            all_usible_article.append(str(i))
    if count_of_articles == 2:
        for i in range(10, 99):
            all_usible_article.append(str(i))
    if count_of_articles == 3:
        for i in range(100, 999):
            all_usible_article.append(str(i))
    if count_of_articles == 4:
        for i in range(1000, 9999):
            all_usible_article.append(str(i))
    if count_of_articles == 5:
        for i in range(10000, 99999):
            all_usible_article.append(str(i))
    if count_of_articles == 6:
        for i in range(100000, 999999):
            all_usible_article.append(str(i))
    if count_of_articles == 7:
        for i in range(1000000, 9999999):
            all_usible_article.append(str(i))
    if count_of_articles == 8:
        for i in range(10000000, 99999999):
            all_usible_article.append(str(i))
    return all_usible_article


def main():
    count_of_articles = 4
    count_with_hyphen = False
    all_use_article = make_all_visible_articles(count_of_articles)
    art = work_doc_csv('./Таблица только артикулов.xlsx')
    list = art.open_File()
    sort = sorted_article_csv()
    sorted_use_article = sort.view_all_article(list, count_of_articles, count_with_hyphen)
    funk = function_with_article(all_use_article)
    free_to_use_article = funk.delete_one_article_from_another(sorted_use_article)
    Data = pd.DataFrame(free_to_use_article)
    if count_with_hyphen == False:
        Data.to_csv('actual articles without hyphen.csv')
    if count_with_hyphen == True:
        Data.to_csv('actual articles with hyphen.csv')
    #art.write_File(count_of_articles, count_with_hyphen, free_to_use_article)

if __name__ == '__main__':
    main()
